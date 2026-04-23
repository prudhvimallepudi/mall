from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, Cookie, Header
from fastapi.responses import JSONResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import random
import uuid
import httpx
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone, timedelta

#from emergentintegrations.llm.chat import LlmChat, UserMessage

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

MONGO_URL = os.environ['MONGO_URL']
DB_NAME = os.environ['DB_NAME']
EMERGENT_LLM_KEY = os.environ.get('EMERGENT_LLM_KEY')

client = AsyncIOMotorClient(MONGO_URL)
db = client[DB_NAME]

app = FastAPI(title="Restaurant Management Dashboard API")
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# =================== MODELS ===================
class User(BaseModel):
    user_id: str
    email: str
    name: str
    picture: Optional[str] = None
    avatar_id: Optional[str] = None
    business_name: Optional[str] = None
    gst_number: Optional[str] = None
    logo_url: Optional[str] = None
    created_at: datetime


class ProfileUpdateRequest(BaseModel):
    name: Optional[str] = None
    avatar_id: Optional[str] = None
    business_name: Optional[str] = None
    gst_number: Optional[str] = None
    logo_url: Optional[str] = None

class SessionExchangeRequest(BaseModel):
    session_id: str

class Branch(BaseModel):
    branch_id: str
    user_id: str
    name: str
    location: str

class InventoryItem(BaseModel):
    item_id: str
    user_id: str
    branch_id: str
    name: str
    unit: str
    stock: float
    min_stock: float
    cost_per_unit: float
    updated_at: datetime

class Expense(BaseModel):
    expense_id: str
    user_id: str
    branch_id: str
    category: str
    amount: float
    note: str
    date: str

class MenuItem(BaseModel):
    item_id: str
    user_id: str
    branch_id: str
    name: str
    category: str
    cost: float
    price: float
    units_sold: int
    waste_units: int

class AddExpenseRequest(BaseModel):
    category: str
    amount: float
    note: str = ""
    branch_id: Optional[str] = None

class AddInventoryRequest(BaseModel):
    name: str
    unit: str
    stock: float
    min_stock: float
    cost_per_unit: float
    branch_id: Optional[str] = None

class UpdateStockRequest(BaseModel):
    delta: float

# =================== AUTH HELPERS ===================
async def get_current_user(
    request: Request,
    session_token: Optional[str] = Cookie(default=None),
    authorization: Optional[str] = Header(default=None),
) -> User:
    token = session_token
    if not token and authorization and authorization.lower().startswith("bearer "):
        token = authorization.split(" ", 1)[1].strip()
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")

    session_doc = await db.user_sessions.find_one({"session_token": token}, {"_id": 0})
    if not session_doc:
        raise HTTPException(status_code=401, detail="Invalid session")

    expires_at = session_doc["expires_at"]
    if isinstance(expires_at, str):
        expires_at = datetime.fromisoformat(expires_at)
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    if expires_at < datetime.now(timezone.utc):
        raise HTTPException(status_code=401, detail="Session expired")

    user_doc = await db.users.find_one({"user_id": session_doc["user_id"]}, {"_id": 0})
    if not user_doc:
        raise HTTPException(status_code=401, detail="User not found")
    return User(**user_doc)


# =================== DEMO DATA SEEDER ===================
DEMO_CATEGORIES = ["Rent", "Electricity", "Gas", "Salaries", "Vendors", "Misc"]
DEMO_MENU = [
    ("Butter Chicken", "Main", 180, 420),
    ("Paneer Tikka", "Starter", 90, 260),
    ("Veg Biryani", "Main", 120, 320),
    ("Masala Dosa", "Breakfast", 45, 140),
    ("Margherita Pizza", "Main", 130, 380),
    ("Caesar Salad", "Starter", 80, 240),
    ("Chocolate Lava Cake", "Dessert", 60, 220),
    ("Cold Coffee", "Beverage", 35, 160),
    ("Gulab Jamun", "Dessert", 25, 90),
    ("Chicken Fried Rice", "Main", 110, 290),
]
DEMO_INVENTORY = [
    ("Chicken", "kg", 12.5, 5, 240),
    ("Paneer", "kg", 4.2, 3, 320),
    ("Basmati Rice", "kg", 28.0, 10, 90),
    ("Tomato", "kg", 15.0, 8, 40),
    ("Onion", "kg", 22.0, 10, 35),
    ("Refined Oil", "L", 9.5, 15, 140),
    ("Flour", "kg", 18.0, 10, 45),
    ("Mozzarella", "kg", 2.1, 4, 520),
    ("Milk", "L", 16.0, 10, 60),
    ("Sugar", "kg", 11.0, 5, 48),
]


async def seed_user_data(user_id: str):
    """Seed demo branches, inventory, expenses, menu, sales for a new user."""
    # Branches
    branches = [
        {"branch_id": f"br_{uuid.uuid4().hex[:8]}", "user_id": user_id, "name": "Downtown Flagship", "location": "Mumbai"},
        {"branch_id": f"br_{uuid.uuid4().hex[:8]}", "user_id": user_id, "name": "Bandra Branch", "location": "Mumbai"},
        {"branch_id": f"br_{uuid.uuid4().hex[:8]}", "user_id": user_id, "name": "Koregaon Park", "location": "Pune"},
    ]
    await db.branches.insert_many([dict(b) for b in branches])

    now = datetime.now(timezone.utc)

    # Menu items per branch
    menu_docs = []
    for b in branches:
        for name, cat, cost, price in DEMO_MENU:
            menu_docs.append({
                "item_id": f"mi_{uuid.uuid4().hex[:8]}",
                "user_id": user_id,
                "branch_id": b["branch_id"],
                "name": name,
                "category": cat,
                "cost": cost,
                "price": price,
                "units_sold": random.randint(40, 400),
                "waste_units": random.randint(0, 12),
            })
    await db.menu_items.insert_many(menu_docs)

    # Inventory per branch
    inv_docs = []
    for b in branches:
        for name, unit, stock, min_stock, cpu in DEMO_INVENTORY:
            # Randomize some items to be low stock
            randomized_stock = stock * random.uniform(0.3, 1.4)
            inv_docs.append({
                "item_id": f"inv_{uuid.uuid4().hex[:8]}",
                "user_id": user_id,
                "branch_id": b["branch_id"],
                "name": name,
                "unit": unit,
                "stock": round(randomized_stock, 2),
                "min_stock": min_stock,
                "cost_per_unit": cpu,
                "updated_at": now,
            })
    await db.inventory.insert_many(inv_docs)

    # Expenses last 60 days
    exp_docs = []
    for b in branches:
        for d in range(60):
            date = (now - timedelta(days=d)).date().isoformat()
            for cat in random.sample(DEMO_CATEGORIES, k=random.randint(1, 3)):
                exp_docs.append({
                    "expense_id": f"ex_{uuid.uuid4().hex[:8]}",
                    "user_id": user_id,
                    "branch_id": b["branch_id"],
                    "category": cat,
                    "amount": round(random.uniform(500, 12000), 2),
                    "note": f"Auto {cat.lower()} entry",
                    "date": date,
                })
    await db.expenses.insert_many(exp_docs)

    # Daily sales aggregates last 90 days
    sales_docs = []
    for b in branches:
        base = random.uniform(18000, 35000)
        for d in range(90):
            date = (now - timedelta(days=d)).date().isoformat()
            total = base * random.uniform(0.7, 1.35)
            orders = int(total / random.uniform(280, 420))
            cash = total * random.uniform(0.1, 0.25)
            upi = total * random.uniform(0.4, 0.6)
            card = total - cash - upi
            sales_docs.append({
                "sale_id": f"s_{uuid.uuid4().hex[:10]}",
                "user_id": user_id,
                "branch_id": b["branch_id"],
                "date": date,
                "total": round(total, 2),
                "orders": orders,
                "cash": round(max(cash, 0), 2),
                "upi": round(max(upi, 0), 2),
                "card": round(max(card, 0), 2),
                "tax": round(total * 0.05, 2),
            })
    await db.sales.insert_many(sales_docs)

    # Employees + 30 days attendance
    await seed_employees(user_id, branches)


# =================== AUTH ROUTES ===================
@api_router.post("/auth/session")
async def auth_session(req: SessionExchangeRequest, response: Response):
    """Exchange session_id from Emergent Auth for a session_token."""
    async with httpx.AsyncClient(timeout=15.0) as http:
        r = await http.get(
            "https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data",
            headers={"X-Session-ID": req.session_id},
        )
        if r.status_code != 200:
            raise HTTPException(status_code=401, detail="Invalid session_id")
        data = r.json()

    email = data["email"]
    existing = await db.users.find_one({"email": email}, {"_id": 0})
    if existing:
        user_id = existing["user_id"]
        await db.users.update_one(
            {"user_id": user_id},
            {"$set": {"name": data.get("name", existing.get("name")), "picture": data.get("picture", existing.get("picture"))}},
        )
    else:
        user_id = f"user_{uuid.uuid4().hex[:12]}"
        await db.users.insert_one({
            "user_id": user_id,
            "email": email,
            "name": data.get("name", email),
            "picture": data.get("picture"),
            "created_at": datetime.now(timezone.utc),
        })
        # Seed demo data for new user
        await seed_user_data(user_id)

    session_token = data["session_token"]
    expires_at = datetime.now(timezone.utc) + timedelta(days=7)
    await db.user_sessions.insert_one({
        "user_id": user_id,
        "session_token": session_token,
        "expires_at": expires_at,
        "created_at": datetime.now(timezone.utc),
    })

    response.set_cookie(
        key="session_token",
        value=session_token,
        httponly=True,
        secure=True,
        samesite="none",
        max_age=7 * 24 * 60 * 60,
        path="/",
    )
    user_doc = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    return {"user": user_doc, "session_token": session_token}


@api_router.get("/auth/me")
async def auth_me(request: Request,
                  session_token: Optional[str] = Cookie(default=None),
                  authorization: Optional[str] = Header(default=None)):
    user = await get_current_user(request, session_token, authorization)
    return user.model_dump()


@api_router.post("/auth/logout")
async def logout(response: Response,
                 session_token: Optional[str] = Cookie(default=None),
                 authorization: Optional[str] = Header(default=None)):
    token = session_token
    if not token and authorization and authorization.lower().startswith("bearer "):
        token = authorization.split(" ", 1)[1].strip()
    if token:
        await db.user_sessions.delete_one({"session_token": token})
    response.delete_cookie("session_token", path="/")
    return {"ok": True}


# =================== DEMO LOGIN (for testing) ===================
class DemoLoginRequest(BaseModel):
    email: str
    name: str = "Demo Owner"


@api_router.post("/auth/demo-login")
async def demo_login(req: DemoLoginRequest, response: Response):
    """Dev/demo-only email login that skips Google. Creates user + session and seeds demo data."""
    existing = await db.users.find_one({"email": req.email}, {"_id": 0})
    if existing:
        user_id = existing["user_id"]
    else:
        user_id = f"user_{uuid.uuid4().hex[:12]}"
        await db.users.insert_one({
            "user_id": user_id,
            "email": req.email,
            "name": req.name,
            "picture": None,
            "created_at": datetime.now(timezone.utc),
        })
        await seed_user_data(user_id)

    session_token = f"demo_{uuid.uuid4().hex}"
    expires_at = datetime.now(timezone.utc) + timedelta(days=7)
    await db.user_sessions.insert_one({
        "user_id": user_id,
        "session_token": session_token,
        "expires_at": expires_at,
        "created_at": datetime.now(timezone.utc),
    })
    response.set_cookie(
        key="session_token", value=session_token, httponly=True,
        secure=True, samesite="none", max_age=7 * 24 * 60 * 60, path="/",
    )
    user_doc = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    return {"user": user_doc, "session_token": session_token}


# =================== EMPLOYEES ===================
class EmployeeCreate(BaseModel):
    name: str
    role: str
    branch_id: Optional[str] = None
    monthly_salary: float
    phone: Optional[str] = ""
    shift: str = "Morning"  # Morning | Evening | Night


class AttendanceRequest(BaseModel):
    employee_id: str
    status: str  # present | absent | leave | half
    date: Optional[str] = None  # YYYY-MM-DD


DEMO_EMPLOYEES = [
    ("Rahul Sharma", "Head Chef", 55000, "Morning"),
    ("Priya Verma", "Sous Chef", 38000, "Morning"),
    ("Ankit Patel", "Waiter", 18000, "Evening"),
    ("Neha Gupta", "Cashier", 22000, "Morning"),
    ("Vikram Singh", "Kitchen Helper", 16000, "Night"),
    ("Deepa Iyer", "Manager", 48000, "Morning"),
    ("Arjun Rao", "Waiter", 18000, "Evening"),
    ("Sana Khan", "Barista", 20000, "Morning"),
]


async def seed_employees(user_id: str, branches: List[Dict[str, Any]]):
    docs = []
    now = datetime.now(timezone.utc)
    for b in branches:
        for name, role, salary, shift in DEMO_EMPLOYEES:
            emp_id = f"emp_{uuid.uuid4().hex[:8]}"
            docs.append({
                "employee_id": emp_id,
                "user_id": user_id,
                "branch_id": b["branch_id"],
                "name": name,
                "role": role,
                "phone": f"+91 9{random.randint(100000000, 999999999)}",
                "monthly_salary": salary,
                "shift": shift,
                "joined": (now - timedelta(days=random.randint(60, 720))).date().isoformat(),
            })
    if docs:
        await db.employees.insert_many([dict(d) for d in docs])
        # Seed 30 days of attendance for each employee
        att_docs = []
        for d in docs:
            for offset in range(30):
                date_str = (now - timedelta(days=offset)).date().isoformat()
                # 85% present, 5% leave, 5% half, 5% absent
                rnd = random.random()
                status = "present" if rnd < 0.85 else "leave" if rnd < 0.9 else "half" if rnd < 0.95 else "absent"
                att_docs.append({
                    "employee_id": d["employee_id"],
                    "user_id": user_id,
                    "branch_id": d["branch_id"],
                    "date": date_str,
                    "status": status,
                })
        if att_docs:
            await db.attendance.insert_many(att_docs)


@api_router.get("/employees")
async def list_employees(request: Request, branch_id: Optional[str] = None,
                        session_token: Optional[str] = Cookie(default=None),
                        authorization: Optional[str] = Header(default=None)):
    user = await get_current_user(request, session_token, authorization)
    q = {"user_id": user.user_id}
    if branch_id and branch_id != "all":
        q["branch_id"] = branch_id
    employees = await db.employees.find(q, {"_id": 0}).to_list(500)

    # Get today's attendance status for each
    today = datetime.now(timezone.utc).date().isoformat()
    emp_ids = [e["employee_id"] for e in employees]
    today_att = await db.attendance.find(
        {"user_id": user.user_id, "date": today, "employee_id": {"$in": emp_ids}},
        {"_id": 0}
    ).to_list(1000)
    att_map = {a["employee_id"]: a["status"] for a in today_att}

    for e in employees:
        e["today_status"] = att_map.get(e["employee_id"], "unmarked")

    total_salary = sum(e["monthly_salary"] for e in employees)
    present_today = sum(1 for e in employees if e["today_status"] == "present")

    return {
        "employees": sorted(employees, key=lambda x: x["name"]),
        "total_employees": len(employees),
        "total_monthly_salary": total_salary,
        "present_today": present_today,
    }


@api_router.post("/employees")
async def add_employee(req: EmployeeCreate, request: Request,
                       session_token: Optional[str] = Cookie(default=None),
                       authorization: Optional[str] = Header(default=None)):
    user = await get_current_user(request, session_token, authorization)
    branch_id = req.branch_id
    if not branch_id:
        first = await db.branches.find_one({"user_id": user.user_id}, {"_id": 0})
        if not first:
            raise HTTPException(status_code=400, detail="No branch available")
        branch_id = first["branch_id"]
    doc = {
        "employee_id": f"emp_{uuid.uuid4().hex[:8]}",
        "user_id": user.user_id,
        "branch_id": branch_id,
        "name": req.name,
        "role": req.role,
        "phone": req.phone or "",
        "monthly_salary": req.monthly_salary,
        "shift": req.shift,
        "joined": datetime.now(timezone.utc).date().isoformat(),
    }
    await db.employees.insert_one(dict(doc))
    return doc


@api_router.post("/employees/attendance")
async def mark_attendance(req: AttendanceRequest, request: Request,
                          session_token: Optional[str] = Cookie(default=None),
                          authorization: Optional[str] = Header(default=None)):
    user = await get_current_user(request, session_token, authorization)
    if req.status not in ("present", "absent", "leave", "half", "unmarked"):
        raise HTTPException(status_code=400, detail="Invalid status")
    date = req.date or datetime.now(timezone.utc).date().isoformat()
    emp = await db.employees.find_one({"employee_id": req.employee_id, "user_id": user.user_id}, {"_id": 0})
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")

    if req.status == "unmarked":
        await db.attendance.delete_one({"user_id": user.user_id, "employee_id": req.employee_id, "date": date})
    else:
        await db.attendance.update_one(
            {"user_id": user.user_id, "employee_id": req.employee_id, "date": date},
            {"$set": {
                "user_id": user.user_id, "employee_id": req.employee_id,
                "branch_id": emp["branch_id"], "date": date, "status": req.status,
            }},
            upsert=True,
        )
    return {"ok": True, "employee_id": req.employee_id, "status": req.status, "date": date}


@api_router.get("/employees/{employee_id}/summary")
async def employee_summary(employee_id: str, request: Request,
                           session_token: Optional[str] = Cookie(default=None),
                           authorization: Optional[str] = Header(default=None)):
    user = await get_current_user(request, session_token, authorization)
    emp = await db.employees.find_one({"employee_id": employee_id, "user_id": user.user_id}, {"_id": 0})
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")

    # Last 30 days attendance
    cutoff = (datetime.now(timezone.utc) - timedelta(days=30)).date().isoformat()
    att = await db.attendance.find(
        {"user_id": user.user_id, "employee_id": employee_id, "date": {"$gte": cutoff}},
        {"_id": 0},
    ).sort("date", -1).to_list(100)

    counts = {"present": 0, "absent": 0, "leave": 0, "half": 0}
    for a in att:
        counts[a["status"]] = counts.get(a["status"], 0) + 1

    working_days = counts["present"] + counts["half"] * 0.5
    estimated_pay = round((emp["monthly_salary"] / 30) * working_days, 2)

    return {
        "employee": emp,
        "attendance_30d": att,
        "counts_30d": counts,
        "estimated_pay_30d": estimated_pay,
    }


# =================== PROFILE ===================
@api_router.patch("/profile")
async def update_profile(req: ProfileUpdateRequest, request: Request,
                         session_token: Optional[str] = Cookie(default=None),
                         authorization: Optional[str] = Header(default=None)):
    user = await get_current_user(request, session_token, authorization)
    updates = {k: v for k, v in req.model_dump().items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="No fields to update")
    await db.users.update_one({"user_id": user.user_id}, {"$set": updates})
    user_doc = await db.users.find_one({"user_id": user.user_id}, {"_id": 0})
    return user_doc


# =================== BRANCHES ===================
@api_router.get("/branches")
async def list_branches(request: Request,
                        session_token: Optional[str] = Cookie(default=None),
                        authorization: Optional[str] = Header(default=None)):
    user = await get_current_user(request, session_token, authorization)
    docs = await db.branches.find({"user_id": user.user_id}, {"_id": 0}).to_list(100)
    return docs


# =================== DASHBOARD ===================
@api_router.get("/dashboard/summary")
async def dashboard_summary(request: Request, branch_id: Optional[str] = None,
                            session_token: Optional[str] = Cookie(default=None),
                            authorization: Optional[str] = Header(default=None)):
    user = await get_current_user(request, session_token, authorization)
    q = {"user_id": user.user_id}
    if branch_id and branch_id != "all":
        q["branch_id"] = branch_id

    today = datetime.now(timezone.utc).date().isoformat()
    yesterday = (datetime.now(timezone.utc) - timedelta(days=1)).date().isoformat()

    # Today KPIs
    today_sales = await db.sales.find({**q, "date": today}, {"_id": 0}).to_list(100)
    yesterday_sales = await db.sales.find({**q, "date": yesterday}, {"_id": 0}).to_list(100)
    total_today = sum(s["total"] for s in today_sales)
    total_yesterday = sum(s["total"] for s in yesterday_sales) or 1
    orders_today = sum(s["orders"] for s in today_sales)
    avg_ticket = (total_today / orders_today) if orders_today else 0

    # 7-day sales trend
    last7 = []
    for i in range(6, -1, -1):
        d = (datetime.now(timezone.utc) - timedelta(days=i)).date().isoformat()
        day_sales = await db.sales.find({**q, "date": d}, {"_id": 0}).to_list(100)
        last7.append({"date": d, "total": round(sum(s["total"] for s in day_sales), 2)})

    # Payment mode breakdown (last 7 days)
    cutoff = (datetime.now(timezone.utc) - timedelta(days=7)).date().isoformat()
    recent = await db.sales.find({**q, "date": {"$gte": cutoff}}, {"_id": 0}).to_list(1000)
    cash = sum(s["cash"] for s in recent)
    upi = sum(s["upi"] for s in recent)
    card = sum(s["card"] for s in recent)

    # Top items
    menu_q = {"user_id": user.user_id}
    if branch_id and branch_id != "all":
        menu_q["branch_id"] = branch_id
    menu = await db.menu_items.find(menu_q, {"_id": 0}).to_list(500)
    agg: Dict[str, Dict[str, Any]] = {}
    for m in menu:
        a = agg.setdefault(m["name"], {"name": m["name"], "units_sold": 0, "revenue": 0})
        a["units_sold"] += m["units_sold"]
        a["revenue"] += m["units_sold"] * m["price"]
    top = sorted(agg.values(), key=lambda x: x["units_sold"], reverse=True)[:5]

    # Low stock count
    inv_q = {"user_id": user.user_id}
    if branch_id and branch_id != "all":
        inv_q["branch_id"] = branch_id
    low_stock = await db.inventory.count_documents({**inv_q, "$expr": {"$lt": ["$stock", "$min_stock"]}})

    # 30-day expense total
    cutoff30 = (datetime.now(timezone.utc) - timedelta(days=30)).date().isoformat()
    exp_recent = await db.expenses.find({**q, "date": {"$gte": cutoff30}}, {"_id": 0}).to_list(5000)
    expenses_30d = sum(e["amount"] for e in exp_recent)

    # 30-day revenue and profit
    sales_30d_docs = await db.sales.find({**q, "date": {"$gte": cutoff30}}, {"_id": 0}).to_list(5000)
    revenue_30d = sum(s["total"] for s in sales_30d_docs)
    profit_30d = revenue_30d - expenses_30d

    change_pct = ((total_today - total_yesterday) / total_yesterday) * 100 if total_yesterday else 0

    return {
        "kpis": {
            "today_sales": round(total_today, 2),
            "today_orders": orders_today,
            "avg_ticket": round(avg_ticket, 2),
            "today_change_pct": round(change_pct, 2),
            "profit_30d": round(profit_30d, 2),
            "revenue_30d": round(revenue_30d, 2),
            "expenses_30d": round(expenses_30d, 2),
            "low_stock_count": low_stock,
        },
        "sales_7d": last7,
        "payment_modes": {"cash": round(cash, 2), "upi": round(upi, 2), "card": round(card, 2)},
        "top_items": top,
    }


@api_router.get("/dashboard/reports")
async def dashboard_reports(request: Request, period: str = "daily", branch_id: Optional[str] = None,
                            session_token: Optional[str] = Cookie(default=None),
                            authorization: Optional[str] = Header(default=None)):
    user = await get_current_user(request, session_token, authorization)
    q = {"user_id": user.user_id}
    if branch_id and branch_id != "all":
        q["branch_id"] = branch_id
    sales = await db.sales.find(q, {"_id": 0}).to_list(5000)

    buckets: Dict[str, float] = {}
    orders_buckets: Dict[str, int] = {}

    def key_for(date_str: str) -> str:
        d = datetime.fromisoformat(date_str)
        if period == "daily":
            return date_str
        if period == "monthly":
            return d.strftime("%Y-%m")
        if period == "yearly":
            return d.strftime("%Y")
        return date_str

    for s in sales:
        k = key_for(s["date"])
        buckets[k] = buckets.get(k, 0) + s["total"]
        orders_buckets[k] = orders_buckets.get(k, 0) + s["orders"]

    if period == "daily":
        keys = sorted(buckets.keys())[-14:]
    elif period == "monthly":
        keys = sorted(buckets.keys())[-6:]
    else:
        keys = sorted(buckets.keys())[-3:]

    return {
        "period": period,
        "points": [{"label": k, "total": round(buckets[k], 2), "orders": orders_buckets[k]} for k in keys],
    }


# =================== INVENTORY ===================
@api_router.get("/inventory")
async def list_inventory(request: Request, branch_id: Optional[str] = None,
                         session_token: Optional[str] = Cookie(default=None),
                         authorization: Optional[str] = Header(default=None)):
    user = await get_current_user(request, session_token, authorization)
    q = {"user_id": user.user_id}
    if branch_id and branch_id != "all":
        q["branch_id"] = branch_id
    items = await db.inventory.find(q, {"_id": 0}).to_list(1000)
    # Sort: low stock first
    items.sort(key=lambda x: (x["stock"] >= x["min_stock"], x["name"]))
    return items


@api_router.post("/inventory")
async def add_inventory(req: AddInventoryRequest, request: Request,
                        session_token: Optional[str] = Cookie(default=None),
                        authorization: Optional[str] = Header(default=None)):
    user = await get_current_user(request, session_token, authorization)
    branch_id = req.branch_id
    if not branch_id:
        first = await db.branches.find_one({"user_id": user.user_id}, {"_id": 0})
        if not first:
            raise HTTPException(status_code=400, detail="No branch available")
        branch_id = first["branch_id"]
    doc = {
        "item_id": f"inv_{uuid.uuid4().hex[:8]}",
        "user_id": user.user_id,
        "branch_id": branch_id,
        "name": req.name,
        "unit": req.unit,
        "stock": req.stock,
        "min_stock": req.min_stock,
        "cost_per_unit": req.cost_per_unit,
        "updated_at": datetime.now(timezone.utc),
    }
    await db.inventory.insert_one(dict(doc))
    return {k: v for k, v in doc.items()}


@api_router.patch("/inventory/{item_id}/stock")
async def update_stock(item_id: str, req: UpdateStockRequest, request: Request,
                       session_token: Optional[str] = Cookie(default=None),
                       authorization: Optional[str] = Header(default=None)):
    user = await get_current_user(request, session_token, authorization)
    res = await db.inventory.find_one_and_update(
        {"item_id": item_id, "user_id": user.user_id},
        {"$inc": {"stock": req.delta}, "$set": {"updated_at": datetime.now(timezone.utc)}},
        projection={"_id": 0}, return_document=True,
    )
    if not res:
        raise HTTPException(status_code=404, detail="Item not found")
    return res


# =================== EXPENSES ===================
@api_router.get("/expenses")
async def list_expenses(request: Request, branch_id: Optional[str] = None, days: int = 30,
                        session_token: Optional[str] = Cookie(default=None),
                        authorization: Optional[str] = Header(default=None)):
    user = await get_current_user(request, session_token, authorization)
    cutoff = (datetime.now(timezone.utc) - timedelta(days=days)).date().isoformat()
    q = {"user_id": user.user_id, "date": {"$gte": cutoff}}
    if branch_id and branch_id != "all":
        q["branch_id"] = branch_id
    items = await db.expenses.find(q, {"_id": 0}).sort("date", -1).to_list(5000)

    # Category breakdown
    cats: Dict[str, float] = {}
    for e in items:
        cats[e["category"]] = cats.get(e["category"], 0) + e["amount"]

    return {
        "items": items[:200],
        "total": round(sum(e["amount"] for e in items), 2),
        "categories": [{"category": k, "amount": round(v, 2)} for k, v in sorted(cats.items(), key=lambda x: -x[1])],
    }


@api_router.post("/expenses")
async def add_expense(req: AddExpenseRequest, request: Request,
                      session_token: Optional[str] = Cookie(default=None),
                      authorization: Optional[str] = Header(default=None)):
    user = await get_current_user(request, session_token, authorization)
    branch_id = req.branch_id
    if not branch_id:
        first = await db.branches.find_one({"user_id": user.user_id}, {"_id": 0})
        if not first:
            raise HTTPException(status_code=400, detail="No branch available")
        branch_id = first["branch_id"]
    doc = {
        "expense_id": f"ex_{uuid.uuid4().hex[:8]}",
        "user_id": user.user_id,
        "branch_id": branch_id,
        "category": req.category,
        "amount": req.amount,
        "note": req.note,
        "date": datetime.now(timezone.utc).date().isoformat(),
    }
    await db.expenses.insert_one(dict(doc))
    return doc


# =================== MENU ANALYTICS ===================
@api_router.get("/menu/analytics")
async def menu_analytics(request: Request, branch_id: Optional[str] = None,
                         session_token: Optional[str] = Cookie(default=None),
                         authorization: Optional[str] = Header(default=None)):
    user = await get_current_user(request, session_token, authorization)
    q = {"user_id": user.user_id}
    if branch_id and branch_id != "all":
        q["branch_id"] = branch_id
    items = await db.menu_items.find(q, {"_id": 0}).to_list(1000)

    agg: Dict[str, Dict[str, Any]] = {}
    for m in items:
        a = agg.setdefault(m["name"], {
            "name": m["name"], "category": m["category"], "cost": m["cost"],
            "price": m["price"], "units_sold": 0, "waste_units": 0,
        })
        a["units_sold"] += m["units_sold"]
        a["waste_units"] += m["waste_units"]
    for a in agg.values():
        a["revenue"] = round(a["units_sold"] * a["price"], 2)
        a["profit"] = round((a["price"] - a["cost"]) * a["units_sold"], 2)
        a["margin_pct"] = round(((a["price"] - a["cost"]) / a["price"]) * 100, 1) if a["price"] else 0

    out = sorted(agg.values(), key=lambda x: -x["revenue"])
    return {"items": out}


# =================== AI INSIGHTS ===================
@api_router.get("/ai/insights")
async def ai_insights(request: Request, branch_id: Optional[str] = None,
                      session_token: Optional[str] = Cookie(default=None),
                      authorization: Optional[str] = Header(default=None)):
    user = await get_current_user(request, session_token, authorization)

    # Build a compact payload of the last 30 days for the LLM
    summary = await dashboard_summary(request, branch_id, session_token, authorization)
    menu = await menu_analytics(request, branch_id, session_token, authorization)
    expenses = await list_expenses(request, branch_id, 30, session_token, authorization)

    top5 = menu["items"][:5]
    bottom5 = sorted(menu["items"], key=lambda x: x["revenue"])[:5]

    payload = {
        "kpis": summary["kpis"],
        "last_7_days_sales": summary["sales_7d"],
        "payment_modes": summary["payment_modes"],
        "top_5_menu_by_revenue": [{"name": x["name"], "revenue": x["revenue"], "margin_pct": x["margin_pct"]} for x in top5],
        "bottom_5_menu": [{"name": x["name"], "revenue": x["revenue"]} for x in bottom5],
        "expense_categories": expenses["categories"],
    }

    system = (
        "You are a senior restaurant business analyst. Given aggregated sales, menu, and expense data, "
        "produce 4 SHORT insight cards. Return strictly valid JSON: "
        "{\"insights\": [{\"type\": one of [prediction, recommendation, alert, opportunity], "
        "\"title\": string, \"message\": string, \"impact\": string}]} "
        "Keep each message under 220 characters, actionable, specific, using numbers when available. No markdown."
    )

    #try:
        #chat = LlmChat(
            #api_key=EMERGENT_LLM_KEY,
            #session_id=f"insights_{user.user_id}_{uuid.uuid4().hex[:6]}",
            #system_message=system,
        #).with_model("anthropic", "claude-sonnet-4-5-20250929")

        #msg = UserMessage(text=f"Data:\n{payload}\n\nReturn only JSON.")
        #raw = await chat.send_message(msg)

        #import json as _json
        #text = raw if isinstance(raw, str) else str(raw)
        # Extract JSON
        #start = text.find("{")
        #end = text.rfind("}")
        #if start != -1 and end != -1:
            #text = text[start:end + 1]
        #parsed = _json.loads(text)
        #insights = parsed.get("insights", [])[:4]
    #except Exception as e:
        #logger.warning(f"LLM insight fallback: {e}")
        insights = [
            {"type": "prediction", "title": "Sales Forecast",
             "message": f"Based on 7-day trend, expect ~₹{int(sum(p['total'] for p in summary['sales_7d']) / 7):,} avg daily revenue next week.",
             "impact": "medium"},
            {"type": "recommendation", "title": "Promote Top Item",
             "message": f"{top5[0]['name'] if top5 else 'Your best seller'} drives highest revenue — feature it in combos to lift AOV.",
             "impact": "high"},
            {"type": "alert", "title": "Low Stock Items",
             "message": f"{summary['kpis']['low_stock_count']} inventory items below threshold. Reorder before weekend.",
             "impact": "high"},
            {"type": "opportunity", "title": "Expense Review",
             "message": f"Top expense: {expenses['categories'][0]['category'] if expenses['categories'] else 'N/A'}. Negotiate vendor contracts to save 8-12%.",
             "impact": "medium"},
        ]

    return {"insights": insights, "generated_at": datetime.now(timezone.utc).isoformat()}


# =================== AI ===================
class AiAskRequest(BaseModel):
    context: str  # dashboard | inventory | expenses | staff | menu
    question: str
    branch_id: Optional[str] = None


CONTEXT_PROMPTS = {
    "dashboard": "You are an executive restaurant advisor. Use the sales & KPI data to answer concisely (<120 words) with numbers.",
    "inventory": "You are a restaurant inventory manager. Focus on low-stock, food cost and reorder priority. <120 words.",
    "expenses": "You are a restaurant CFO. Identify biggest cost buckets and saving opportunities. <120 words.",
    "staff": "You are a restaurant HR manager. Analyse attendance, shifts, payroll. <120 words.",
    "menu": "You are a restaurant menu consultant. Recommend pricing, promotion, margin optimisation. <120 words.",
}

SUGGESTED_QUESTIONS = {
    "dashboard": [
        "Summarise today's sales performance",
        "Which payment mode dominates?",
        "What is my 30-day profit trend?",
    ],
    "inventory": [
        "Which items are critically low?",
        "What should I reorder first?",
        "Estimate my inventory value",
    ],
    "expenses": [
        "What is my biggest cost this month?",
        "Where can I cut 10% expenses?",
        "Which category is rising the fastest?",
    ],
    "staff": [
        "Who has the highest attendance?",
        "Estimate this month's payroll liability",
        "Any attendance red-flags?",
    ],
    "menu": [
        "Which items have the best margin?",
        "What combos should I promote?",
        "Which items are under-performing?",
    ],
}


async def _context_payload(user_id: str, context: str, branch_id: Optional[str], request: Request,
                           session_token: Optional[str], authorization: Optional[str]) -> Dict[str, Any]:
    """Build a small, relevant data snapshot for the AI context."""
    q = {"user_id": user_id}
    if branch_id and branch_id != "all":
        q["branch_id"] = branch_id

    if context == "dashboard":
        summary = await dashboard_summary(request, branch_id, session_token, authorization)
        return {"kpis": summary["kpis"], "sales_7d": summary["sales_7d"],
                "payment_modes": summary["payment_modes"], "top_items": summary["top_items"]}

    if context == "inventory":
        items = await db.inventory.find(q, {"_id": 0}).to_list(500)
        low = [i for i in items if i["stock"] < i["min_stock"]]
        total_value = sum(i["stock"] * i["cost_per_unit"] for i in items)
        return {
            "total_items": len(items),
            "low_stock": [{"name": i["name"], "stock": i["stock"], "min_stock": i["min_stock"], "unit": i["unit"]} for i in low[:20]],
            "top_value": sorted(
                [{"name": i["name"], "value": round(i["stock"] * i["cost_per_unit"], 2)} for i in items],
                key=lambda x: -x["value"],
            )[:10],
            "total_inventory_value": round(total_value, 2),
        }

    if context == "expenses":
        data = await list_expenses(request, branch_id, 30, session_token, authorization)
        return {"total_30d": data["total"], "categories": data["categories"], "recent": data["items"][:10]}

    if context == "staff":
        emp = await list_employees(request, branch_id, session_token, authorization)
        return {
            "total_employees": emp["total_employees"],
            "present_today": emp["present_today"],
            "total_monthly_salary": emp["total_monthly_salary"],
            "by_shift": {
                s: sum(1 for e in emp["employees"] if e["shift"] == s)
                for s in ("Morning", "Evening", "Night")
            },
        }

    if context == "menu":
        m = await menu_analytics(request, branch_id, session_token, authorization)
        return {
            "top_revenue": m["items"][:5],
            "low_revenue": sorted(m["items"], key=lambda x: x["revenue"])[:5],
            "high_margin": sorted(m["items"], key=lambda x: -x["margin_pct"])[:5],
        }
    return {}


@api_router.get("/ai/suggestions")
async def ai_suggestions(context: str = "dashboard"):
    return {"suggestions": SUGGESTED_QUESTIONS.get(context, SUGGESTED_QUESTIONS["dashboard"])}


@api_router.post("/ai/ask")
async def ai_ask(req: AiAskRequest, request: Request,
                 session_token: Optional[str] = Cookie(default=None),
                 authorization: Optional[str] = Header(default=None)):
    user = await get_current_user(request, session_token, authorization)
    ctx = req.context if req.context in CONTEXT_PROMPTS else "dashboard"
    data = await _context_payload(user.user_id, ctx, req.branch_id, request, session_token, authorization)

    system = (
        f"{CONTEXT_PROMPTS[ctx]} Only use the data provided. "
        "Use Indian rupee ₹ formatting. Return plain text (no markdown)."
    )
    #try:
        #chat = LlmChat(
            #api_key=EMERGENT_LLM_KEY,
            #session_id=f"ask_{user.user_id}_{uuid.uuid4().hex[:6]}",
            #system_message=system,
        #).with_model("anthropic", "claude-sonnet-4-5-20250929")
        #msg = UserMessage(text=f"Context ({ctx}) data:\n{data}\n\nUser question: {req.question}")
        #raw = await chat.send_message(msg)
        return {"message": "Feature coming soon"}
        answer = raw if isinstance(raw, str) else str(raw)
    except Exception as e:
        logger.warning(f"AI ask failed: {e}")
        answer = "Sorry, I could not reach the AI service right now. Please try again."

    return {"answer": answer.strip(), "context": ctx}


# =================== CSV IMPORT ===================
class CSVImportRequest(BaseModel):
    type: str  # "sales" | "expenses"
    rows: List[Dict[str, Any]]
    branch_id: Optional[str] = None


@api_router.post("/import/csv")
async def import_csv(req: CSVImportRequest, request: Request,
                     session_token: Optional[str] = Cookie(default=None),
                     authorization: Optional[str] = Header(default=None)):
    user = await get_current_user(request, session_token, authorization)
    branch_id = req.branch_id
    if not branch_id:
        first = await db.branches.find_one({"user_id": user.user_id}, {"_id": 0})
        if not first:
            raise HTTPException(status_code=400, detail="No branch available")
        branch_id = first["branch_id"]

    inserted = 0
    if req.type == "sales":
        docs = []
        for row in req.rows:
            total = float(row.get("total", 0))
            docs.append({
                "sale_id": f"s_{uuid.uuid4().hex[:10]}",
                "user_id": user.user_id,
                "branch_id": branch_id,
                "date": row.get("date", datetime.now(timezone.utc).date().isoformat()),
                "total": total,
                "orders": int(row.get("orders", 1)),
                "cash": float(row.get("cash", total * 0.2)),
                "upi": float(row.get("upi", total * 0.5)),
                "card": float(row.get("card", total * 0.3)),
                "tax": float(row.get("tax", total * 0.05)),
            })
        if docs:
            await db.sales.insert_many(docs)
            inserted = len(docs)
    elif req.type == "expenses":
        docs = []
        for row in req.rows:
            docs.append({
                "expense_id": f"ex_{uuid.uuid4().hex[:8]}",
                "user_id": user.user_id,
                "branch_id": branch_id,
                "category": row.get("category", "Misc"),
                "amount": float(row.get("amount", 0)),
                "note": row.get("note", "CSV import"),
                "date": row.get("date", datetime.now(timezone.utc).date().isoformat()),
            })
        if docs:
            await db.expenses.insert_many(docs)
            inserted = len(docs)
    else:
        raise HTTPException(status_code=400, detail="Unsupported type. Use 'sales' or 'expenses'.")

    return {"ok": True, "inserted": inserted}


@api_router.get("/")
async def root():
    return {"message": "Restaurant Management API", "status": "ok"}


# =================== INTEGRATIONS (File Upload & AI Parsing) ===================
import base64 as _b64
import io as _io
import csv as _csv

TARGET_SCHEMAS = {
    "sales": {
        "fields": ["date", "total", "orders", "cash", "upi", "card", "tax"],
        "required": ["date", "total"],
    },
    "expenses": {
        "fields": ["date", "category", "amount", "note"],
        "required": ["category", "amount"],
    },
    "inventory": {
        "fields": ["name", "unit", "stock", "min_stock", "cost_per_unit"],
        "required": ["name", "stock"],
    },
    "attendance": {
        "fields": ["employee_name", "date", "status"],
        "required": ["employee_name", "status"],
    },
}


class ParseRequest(BaseModel):
    filename: str
    content_base64: str
    mime_type: Optional[str] = None
    category: str  # sales | expenses | inventory | attendance


class ImportRequest(BaseModel):
    upload_id: str
    category: str
    mapping: Dict[str, str]  # target_field -> source_column
    rows: List[Dict[str, Any]]


def _guess_kind(filename: str, mime: Optional[str]) -> str:
    n = (filename or "").lower()
    m = (mime or "").lower()
    if n.endswith(".csv") or "csv" in m:
        return "csv"
    if n.endswith(".xlsx") or n.endswith(".xls") or "spreadsheet" in m or "excel" in m:
        return "excel"
    if n.endswith(".pdf") or "pdf" in m:
        return "pdf"
    if any(n.endswith(ext) for ext in (".png", ".jpg", ".jpeg", ".webp", ".heic")) or m.startswith("image/"):
        return "image"
    if n.endswith(".txt") or "text/plain" in m:
        return "txt"
    return "unknown"


def _parse_csv_bytes(data: bytes) -> tuple[List[str], List[Dict[str, Any]]]:
    text = data.decode("utf-8", errors="ignore")
    reader = _csv.reader(_io.StringIO(text))
    rows = list(reader)
    if not rows:
        return [], []
    headers = [h.strip() for h in rows[0]]
    body = []
    for r in rows[1:]:
        if not any(c.strip() for c in r):
            continue
        body.append({headers[i]: (r[i].strip() if i < len(r) else "") for i in range(len(headers))})
    return headers, body


def _parse_excel_bytes(data: bytes) -> tuple[List[str], List[Dict[str, Any]]]:
    from openpyxl import load_workbook
    wb = load_workbook(_io.BytesIO(data), data_only=True, read_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return [], []
    headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(all_rows[0])]
    body = []
    for r in all_rows[1:]:
        if r is None or all(c is None or str(c).strip() == "" for c in r):
            continue
        body.append({headers[i]: ("" if c is None else str(c).strip()) for i, c in enumerate(r) if i < len(headers)})
    return headers, body


def _parse_pdf_bytes(data: bytes) -> tuple[List[str], List[Dict[str, Any]]]:
    from pypdf import PdfReader
    reader = PdfReader(_io.BytesIO(data))
    text_lines: List[str] = []
    for page in reader.pages[:20]:
        t = page.extract_text() or ""
        for line in t.splitlines():
            line = line.strip()
            if line:
                text_lines.append(line)
    # Heuristic: find a line with commas/tabs that looks like a header, then split
    sep = None
    header_idx = -1
    for i, ln in enumerate(text_lines[:50]):
        if ln.count(",") >= 2:
            sep = ","
            header_idx = i
            break
        if ln.count("\t") >= 2:
            sep = "\t"
            header_idx = i
            break
    if sep is None or header_idx == -1:
        # Return raw text as single-column
        return ["raw_line"], [{"raw_line": l} for l in text_lines]
    headers = [h.strip() for h in text_lines[header_idx].split(sep)]
    body = []
    for ln in text_lines[header_idx + 1:]:
        cells = [c.strip() for c in ln.split(sep)]
        if len(cells) < len(headers):
            continue
        body.append({headers[i]: cells[i] for i in range(len(headers))})
    return headers, body


async def _parse_image_bytes(data: bytes, mime: str, filename: str) -> tuple[List[str], List[Dict[str, Any]]]:
    """Use GPT-4o vision to OCR bill/invoice images and extract tabular data."""
    import json as _json
    #try:
        #from emergentintegrations.llm.chat import LlmChat, UserMessage, ImageContent
        #image_b64 = _b64.b64encode(data).decode()
        #chat = LlmChat(
            #api_key=EMERGENT_LLM_KEY,
            #session_id=f"ocr_{uuid.uuid4().hex[:8]}",
            #system_message=(
                #"You are an OCR & data-extraction expert for restaurant documents "
                #"(bills, invoices, daily sales reports, attendance sheets). "
                #"Extract ANY tabular data you see and return STRICT JSON of the form: "
                #"{\"headers\": [\"col1\", ...], \"rows\": [{\"col1\": \"value\", ...}, ...]}. "
                #"Use clear column names in lowercase_snake_case like date, amount, item_name, qty, total. "
                #"If you see a single-row bill (not a table), still output one row with fields."
            #),
        #).with_model("openai", "gpt-4o")
        #msg = UserMessage(
            #text="Extract all tabular data from this document image.",
            #file_contents=[ImageContent(image_base64=image_b64)],
        #)
        #raw = await chat.send_message(msg)
        return {"message": "Feature coming soon"}
return {
    "headers": [],
    "rows": [],
    "message": "OCR feature temporarily disabled"
}
        text = raw if isinstance(raw, str) else str(raw)
        s, e = text.find("{"), text.rfind("}")
        if s == -1 or e == -1:
            return ["raw_text"], [{"raw_text": text[:500]}]
        parsed = _json.loads(text[s:e + 1])
        headers = parsed.get("headers", []) or []
        rows = parsed.get("rows", []) or []
        # Normalize
        body = []
        for r in rows:
            body.append({h: str(r.get(h, "")) for h in headers})
        return headers, body
    except Exception as ex:
        logger.warning(f"OCR failed: {ex}")
        return ["raw_text"], [{"raw_text": f"OCR error: {ex}"}]


async def _ai_map_columns(headers: List[str], sample_rows: List[Dict[str, Any]], category: str) -> Dict[str, str]:
    """Ask Claude to map source columns to target schema fields."""
    import json as _json
    schema = TARGET_SCHEMAS.get(category)
    if not schema:
        return {}
    #try:
        #chat = LlmChat(
            #api_key=EMERGENT_LLM_KEY,
            #session_id=f"map_{uuid.uuid4().hex[:8]}",
            #system_message=(
                #"You map source spreadsheet/CSV columns to target database fields. "
                #"Return strict JSON: {\"mapping\": {\"target_field\": \"source_column_or_empty_string\"}}. "
                #"Use empty string when no suitable source column exists. Do not invent columns."
            #),
        #).with_model("anthropic", "claude-sonnet-4-5-20250929")
        #prompt = {
            #"target_fields": schema["fields"],
            #"required_fields": schema["required"],
            #"source_headers": headers,
            #"sample_rows": sample_rows[:3],
            #"category": category,
        #}
        #raw = await chat.send_message(UserMessage(text=f"Map columns. Input: {prompt}. Return JSON only."))
        return {"message": "Feature coming soon"}
        return {
    "headers": [],
    "rows": [],
    "message": "OCR feature temporarily disabled"
}
        text = raw if isinstance(raw, str) else str(raw)
        s, e = text.find("{"), text.rfind("}")
        if s == -1:
            return {}
        parsed = _json.loads(text[s:e + 1])
        mapping = parsed.get("mapping", {})
        # Keep only target fields that exist in schema
        return {k: v for k, v in mapping.items() if k in schema["fields"] and isinstance(v, str)}
    except Exception as ex:
        logger.warning(f"AI mapping failed, falling back to heuristic: {ex}")
        # Simple fallback: match by lowercase substring
        mapping = {}
        lower_headers = {h.lower().replace(" ", "_"): h for h in headers}
        for f in schema["fields"]:
            for key, orig in lower_headers.items():
                if f == key or f in key or key in f:
                    mapping[f] = orig
                    break
        return mapping


@api_router.post("/integrations/parse")
async def integrations_parse(req: ParseRequest, request: Request,
                             session_token: Optional[str] = Cookie(default=None),
                             authorization: Optional[str] = Header(default=None)):
    user = await get_current_user(request, session_token, authorization)
    if req.category not in TARGET_SCHEMAS:
        raise HTTPException(status_code=400, detail=f"Unsupported category. Allowed: {list(TARGET_SCHEMAS.keys())}")

    try:
        data = _b64.b64decode(req.content_base64)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid base64 content")
    if len(data) > 5 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="File too large (max 5 MB)")

    kind = _guess_kind(req.filename, req.mime_type)
    errors: List[str] = []

    try:
        if kind == "csv" or kind == "txt":
            headers, rows = _parse_csv_bytes(data)
        elif kind == "excel":
            headers, rows = _parse_excel_bytes(data)
        elif kind == "pdf":
            headers, rows = _parse_pdf_bytes(data)
        elif kind == "image":
            headers, rows = await _parse_image_bytes(data, req.mime_type or "", req.filename)
        else:
            raise HTTPException(status_code=400, detail=f"Unsupported file type: {req.filename}")
    except HTTPException:
        raise
    except Exception as ex:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {ex}")

    # Remove empty rows
    rows = [r for r in rows if any(str(v).strip() for v in r.values())]

    # AI mapping
    mapping = await _ai_map_columns(headers, rows[:3], req.category)

    # Check required fields
    required = TARGET_SCHEMAS[req.category]["required"]
    missing = [f for f in required if not mapping.get(f)]
    if missing:
        errors.append(f"Missing required mappings: {', '.join(missing)}")

    # Dup detection within file (naive)
    seen = set()
    dup_count = 0
    for r in rows:
        key = tuple(sorted(r.items()))
        if key in seen:
            dup_count += 1
        else:
            seen.add(key)

    # Create upload record (metadata only)
    upload_id = f"up_{uuid.uuid4().hex[:10]}"
    await db.uploads.insert_one({
        "upload_id": upload_id,
        "user_id": user.user_id,
        "filename": req.filename,
        "category": req.category,
        "kind": kind,
        "total_rows": len(rows),
        "headers": headers,
        "mapping": mapping,
        "status": "parsed",
        "errors": errors,
        "dup_count": dup_count,
        "imported_count": 0,
        "created_at": datetime.now(timezone.utc),
    })

    return {
        "upload_id": upload_id,
        "filename": req.filename,
        "kind": kind,
        "category": req.category,
        "headers": headers,
        "preview_rows": rows[:10],
        "rows": rows,
        "total_rows": len(rows),
        "ai_mapping": mapping,
        "target_fields": TARGET_SCHEMAS[req.category]["fields"],
        "required_fields": required,
        "errors": errors,
        "dup_count": dup_count,
    }


def _coerce_number(v: Any, default: float = 0.0) -> float:
    if v is None:
        return default
    s = str(v).strip().replace(",", "").replace("₹", "").replace("$", "").replace("Rs.", "").replace("Rs", "")
    try:
        return float(s)
    except Exception:
        return default


def _coerce_date(v: Any) -> str:
    if v is None:
        return datetime.now(timezone.utc).date().isoformat()
    s = str(v).strip()
    if not s:
        return datetime.now(timezone.utc).date().isoformat()
    # Try common formats
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%d %b %Y", "%d %B %Y"):
        try:
            return datetime.strptime(s[:10], fmt).date().isoformat()
        except Exception:
            continue
    return datetime.now(timezone.utc).date().isoformat()


@api_router.post("/integrations/import")
async def integrations_import(req: ImportRequest, request: Request,
                              session_token: Optional[str] = Cookie(default=None),
                              authorization: Optional[str] = Header(default=None)):
    user = await get_current_user(request, session_token, authorization)
    if req.category not in TARGET_SCHEMAS:
        raise HTTPException(status_code=400, detail="Unsupported category")

    first = await db.branches.find_one({"user_id": user.user_id}, {"_id": 0})
    if not first:
        raise HTTPException(status_code=400, detail="No branch available")
    branch_id = first["branch_id"]

    def pick(row: Dict[str, Any], target: str) -> Any:
        col = req.mapping.get(target)
        if not col:
            return None
        return row.get(col)

    inserted = 0
    failed = 0

    if req.category == "sales":
        docs = []
        for r in req.rows:
            try:
                total = _coerce_number(pick(r, "total"))
                if total <= 0:
                    failed += 1
                    continue
                docs.append({
                    "sale_id": f"s_{uuid.uuid4().hex[:10]}",
                    "user_id": user.user_id,
                    "branch_id": branch_id,
                    "date": _coerce_date(pick(r, "date")),
                    "total": total,
                    "orders": int(_coerce_number(pick(r, "orders"), 1)),
                    "cash": _coerce_number(pick(r, "cash"), total * 0.2),
                    "upi": _coerce_number(pick(r, "upi"), total * 0.5),
                    "card": _coerce_number(pick(r, "card"), total * 0.3),
                    "tax": _coerce_number(pick(r, "tax"), total * 0.05),
                })
            except Exception:
                failed += 1
        if docs:
            await db.sales.insert_many(docs)
            inserted = len(docs)

    elif req.category == "expenses":
        docs = []
        for r in req.rows:
            try:
                amount = _coerce_number(pick(r, "amount"))
                cat = str(pick(r, "category") or "Misc").strip() or "Misc"
                if amount <= 0:
                    failed += 1
                    continue
                docs.append({
                    "expense_id": f"ex_{uuid.uuid4().hex[:8]}",
                    "user_id": user.user_id,
                    "branch_id": branch_id,
                    "category": cat,
                    "amount": amount,
                    "note": str(pick(r, "note") or ""),
                    "date": _coerce_date(pick(r, "date")),
                })
            except Exception:
                failed += 1
        if docs:
            await db.expenses.insert_many(docs)
            inserted = len(docs)

    elif req.category == "inventory":
        docs = []
        for r in req.rows:
            try:
                name = str(pick(r, "name") or "").strip()
                stock = _coerce_number(pick(r, "stock"))
                if not name:
                    failed += 1
                    continue
                docs.append({
                    "item_id": f"inv_{uuid.uuid4().hex[:8]}",
                    "user_id": user.user_id,
                    "branch_id": branch_id,
                    "name": name,
                    "unit": str(pick(r, "unit") or "kg"),
                    "stock": stock,
                    "min_stock": _coerce_number(pick(r, "min_stock"), 5),
                    "cost_per_unit": _coerce_number(pick(r, "cost_per_unit")),
                    "updated_at": datetime.now(timezone.utc),
                })
            except Exception:
                failed += 1
        if docs:
            await db.inventory.insert_many(docs)
            inserted = len(docs)

    elif req.category == "attendance":
        att_docs = []
        for r in req.rows:
            try:
                name = str(pick(r, "employee_name") or "").strip()
                status = str(pick(r, "status") or "").strip().lower()
                if status not in ("present", "absent", "leave", "half"):
                    # Common synonyms
                    status_map = {"p": "present", "a": "absent", "l": "leave", "h": "half", "y": "present", "n": "absent"}
                    status = status_map.get(status, "present")
                if not name:
                    failed += 1
                    continue
                # Match employee by name (case-insensitive)
                emp = await db.employees.find_one(
                    {"user_id": user.user_id, "name": {"$regex": f"^{name}$", "$options": "i"}},
                    {"_id": 0},
                )
                if not emp:
                    failed += 1
                    continue
                date_str = _coerce_date(pick(r, "date"))
                await db.attendance.update_one(
                    {"user_id": user.user_id, "employee_id": emp["employee_id"], "date": date_str},
                    {"$set": {
                        "user_id": user.user_id, "employee_id": emp["employee_id"],
                        "branch_id": emp["branch_id"], "date": date_str, "status": status,
                    }},
                    upsert=True,
                )
                att_docs.append(True)
            except Exception:
                failed += 1
        inserted = len(att_docs)

    # Update upload record
    await db.uploads.update_one(
        {"upload_id": req.upload_id, "user_id": user.user_id},
        {"$set": {
            "status": "imported" if inserted > 0 else "failed",
            "imported_count": inserted,
            "failed_count": failed,
            "imported_at": datetime.now(timezone.utc),
        }},
    )

    return {"ok": True, "inserted": inserted, "failed": failed, "upload_id": req.upload_id}


@api_router.get("/integrations/history")
async def integrations_history(request: Request,
                               session_token: Optional[str] = Cookie(default=None),
                               authorization: Optional[str] = Header(default=None)):
    user = await get_current_user(request, session_token, authorization)
    items = await db.uploads.find(
        {"user_id": user.user_id},
        {"_id": 0, "headers": 0, "mapping": 0}
    ).sort("created_at", -1).to_list(50)
    # Convert datetime to iso string
    for it in items:
        for key in ("created_at", "imported_at"):
            if key in it and hasattr(it[key], "isoformat"):
                it[key] = it[key].isoformat()
    return {"items": items}



# =================== NOTIFICATION SETTINGS ===================
DEFAULT_NOTIF_SETTINGS = {
    "low_stock_enabled": True,
    "out_of_stock_enabled": True,
    "staff_absent_enabled": True,
    "high_sales_milestone_enabled": True,
    "low_sales_warning_enabled": True,
    "expense_limit_enabled": True,
    "expense_limit_monthly": 500000.0,  # ₹5L default cap
    "sound_enabled": True,
    "eod_time": "23:30",
    "eod_timezone": "Asia/Kolkata",
    "eod_recipient_emails": [],
    "eod_daily_enabled": True,
    "eod_weekly_enabled": False,
    "eod_monthly_enabled": False,
    "branch_rules": {},  # {branch_id: true/false}
}


class NotifSettingsUpdate(BaseModel):
    low_stock_enabled: Optional[bool] = None
    out_of_stock_enabled: Optional[bool] = None
    staff_absent_enabled: Optional[bool] = None
    high_sales_milestone_enabled: Optional[bool] = None
    low_sales_warning_enabled: Optional[bool] = None
    expense_limit_enabled: Optional[bool] = None
    expense_limit_monthly: Optional[float] = None
    sound_enabled: Optional[bool] = None
    eod_time: Optional[str] = None  # "HH:MM"
    eod_timezone: Optional[str] = None
    eod_recipient_emails: Optional[List[str]] = None
    eod_daily_enabled: Optional[bool] = None
    eod_weekly_enabled: Optional[bool] = None
    eod_monthly_enabled: Optional[bool] = None
    branch_rules: Optional[Dict[str, bool]] = None


async def _get_settings(user_id: str) -> Dict[str, Any]:
    doc = await db.notif_settings.find_one({"user_id": user_id}, {"_id": 0})
    if not doc:
        doc = {"user_id": user_id, **DEFAULT_NOTIF_SETTINGS}
        await db.notif_settings.insert_one(dict(doc))
        doc.pop("_id", None)
    # Backfill any new defaults
    for k, v in DEFAULT_NOTIF_SETTINGS.items():
        doc.setdefault(k, v)
    return doc


@api_router.get("/settings/notifications")
async def get_notif_settings(request: Request,
                             session_token: Optional[str] = Cookie(default=None),
                             authorization: Optional[str] = Header(default=None)):
    user = await get_current_user(request, session_token, authorization)
    settings = await _get_settings(user.user_id)
    settings.pop("user_id", None)
    return settings


@api_router.patch("/settings/notifications")
async def patch_notif_settings(req: NotifSettingsUpdate, request: Request,
                               session_token: Optional[str] = Cookie(default=None),
                               authorization: Optional[str] = Header(default=None)):
    user = await get_current_user(request, session_token, authorization)
    updates = {k: v for k, v in req.model_dump().items() if v is not None}
    # Validate eod_time
    if "eod_time" in updates:
        try:
            hh, mm = updates["eod_time"].split(":")
            if not (0 <= int(hh) < 24 and 0 <= int(mm) < 60):
                raise ValueError()
        except Exception:
            raise HTTPException(status_code=400, detail="eod_time must be HH:MM (24-hour)")
    await _get_settings(user.user_id)  # ensure row exists
    await db.notif_settings.update_one({"user_id": user.user_id}, {"$set": updates})
    settings = await _get_settings(user.user_id)
    settings.pop("user_id", None)
    return settings


# =================== NOTIFICATIONS ENGINE ===================
async def _compute_notifications(user_id: str) -> List[Dict[str, Any]]:
    """Compute current actionable notifications from live DB state."""
    settings = await _get_settings(user_id)
    branch_rules = settings.get("branch_rules") or {}
    allowed_branch = lambda bid: branch_rules.get(bid, True)  # default on

    today = datetime.now(timezone.utc).date().isoformat()
    notifs: List[Dict[str, Any]] = []

    # Branch lookup
    branches = await db.branches.find({"user_id": user_id}, {"_id": 0}).to_list(50)
    branch_name = {b["branch_id"]: b["name"] for b in branches}

    # --- Inventory low / out of stock ---
    if settings.get("low_stock_enabled") or settings.get("out_of_stock_enabled"):
        inv = await db.inventory.find({"user_id": user_id}, {"_id": 0}).to_list(2000)
        for i in inv:
            if not allowed_branch(i["branch_id"]):
                continue
            stock = float(i.get("stock") or 0)
            min_stock = float(i.get("min_stock") or 0)
            bn = branch_name.get(i["branch_id"], "Branch")
            if stock <= 0 and settings.get("out_of_stock_enabled"):
                notifs.append({
                    "id": f"out:{i['item_id']}:{today}",
                    "type": "out_of_stock",
                    "severity": "critical",
                    "icon": "close-circle",
                    "title": f"{i['name']} stock out of stock",
                    "message": f"0 {i['unit']} left at {bn}. Reorder immediately.",
                    "branch_id": i["branch_id"],
                    "branch_name": bn,
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "link": "inventory",
                })
            elif stock < min_stock and settings.get("low_stock_enabled"):
                notifs.append({
                    "id": f"low:{i['item_id']}:{today}",
                    "type": "low_stock",
                    "severity": "warning",
                    "icon": "alert-circle",
                    "title": f"{i['name']} stock is low",
                    "message": f"{stock:g} {i['unit']} remaining at {bn} (min {min_stock:g}).",
                    "branch_id": i["branch_id"],
                    "branch_name": bn,
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "link": "inventory",
                })

    # --- Staff absent today ---
    if settings.get("staff_absent_enabled"):
        absents = await db.attendance.find(
            {"user_id": user_id, "date": today, "status": "absent"}, {"_id": 0}
        ).to_list(500)
        if absents:
            emp_ids = [a["employee_id"] for a in absents]
            emps = await db.employees.find({"user_id": user_id, "employee_id": {"$in": emp_ids}}, {"_id": 0}).to_list(500)
            emp_map = {e["employee_id"]: e for e in emps}
            for a in absents:
                e = emp_map.get(a["employee_id"])
                if not e or not allowed_branch(e["branch_id"]):
                    continue
                bn = branch_name.get(e["branch_id"], "Branch")
                notifs.append({
                    "id": f"absent:{e['employee_id']}:{today}",
                    "type": "staff_absent",
                    "severity": "warning",
                    "icon": "person-remove",
                    "title": f"{e['name']} is absent today",
                    "message": f"{e['role']} · {e['shift']} shift · {bn}. Tap to call or reassign.",
                    "branch_id": e["branch_id"],
                    "branch_name": bn,
                    "employee_id": e["employee_id"],
                    "phone": e.get("phone"),
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "link": "employees",
                })

    # --- Sales milestones ---
    if settings.get("high_sales_milestone_enabled") or settings.get("low_sales_warning_enabled"):
        today_sales_docs = await db.sales.find({"user_id": user_id, "date": today}, {"_id": 0}).to_list(500)
        today_total = sum(s["total"] for s in today_sales_docs)
        cutoff = (datetime.now(timezone.utc) - timedelta(days=7)).date().isoformat()
        recent = await db.sales.find({"user_id": user_id, "date": {"$gte": cutoff}}, {"_id": 0}).to_list(2000)
        # 7-day average per-day (excluding today)
        days = {}
        for s in recent:
            if s["date"] == today:
                continue
            days[s["date"]] = days.get(s["date"], 0) + s["total"]
        avg_7d = (sum(days.values()) / max(len(days), 1)) if days else 0

        if avg_7d > 0:
            if today_total >= avg_7d * 1.5 and settings.get("high_sales_milestone_enabled"):
                notifs.append({
                    "id": f"highsales:{today}",
                    "type": "high_sales",
                    "severity": "success",
                    "icon": "trending-up",
                    "title": "🎉 High sales milestone",
                    "message": f"Today ₹{int(today_total):,} is {int((today_total / avg_7d - 1) * 100)}% above your 7-day average.",
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "link": "dashboard",
                })
            elif today_total > 0 and today_total <= avg_7d * 0.5 and settings.get("low_sales_warning_enabled"):
                notifs.append({
                    "id": f"lowsales:{today}",
                    "type": "low_sales",
                    "severity": "warning",
                    "icon": "trending-down",
                    "title": "Low sales warning",
                    "message": f"Today ₹{int(today_total):,} is {int((1 - today_total / avg_7d) * 100)}% below your 7-day average.",
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "link": "dashboard",
                })

    # --- Monthly expense limit ---
    if settings.get("expense_limit_enabled"):
        limit = float(settings.get("expense_limit_monthly") or 0)
        if limit > 0:
            month_cutoff = datetime.now(timezone.utc).replace(day=1).date().isoformat()
            exps = await db.expenses.find(
                {"user_id": user_id, "date": {"$gte": month_cutoff}}, {"_id": 0}
            ).to_list(5000)
            total = sum(e["amount"] for e in exps)
            if total >= limit:
                notifs.append({
                    "id": f"explimit:{datetime.now(timezone.utc).strftime('%Y-%m')}",
                    "type": "expense_limit",
                    "severity": "critical",
                    "icon": "cash-outline",
                    "title": "Monthly expense limit crossed",
                    "message": f"₹{int(total):,} spent vs ₹{int(limit):,} limit ({int(total / limit * 100)}%).",
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "link": "expenses",
                })
            elif total >= limit * 0.85:
                notifs.append({
                    "id": f"expwarn:{datetime.now(timezone.utc).strftime('%Y-%m')}",
                    "type": "expense_warning",
                    "severity": "warning",
                    "icon": "warning-outline",
                    "title": "Approaching expense limit",
                    "message": f"₹{int(total):,} of ₹{int(limit):,} monthly budget used ({int(total / limit * 100)}%).",
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "link": "expenses",
                })

    # --- Recent imports ---
    recent_imports = await db.uploads.find(
        {"user_id": user_id, "status": "imported"}, {"_id": 0}
    ).sort("imported_at", -1).to_list(3)
    for up in recent_imports:
        ts = up.get("imported_at")
        if isinstance(ts, datetime):
            age = (datetime.now(timezone.utc) - ts.replace(tzinfo=timezone.utc) if ts.tzinfo is None else datetime.now(timezone.utc) - ts).total_seconds()
        else:
            age = 999999
        if age < 24 * 3600:  # 1 day
            notifs.append({
                "id": f"import:{up['upload_id']}",
                "type": "import_complete",
                "severity": "info",
                "icon": "checkmark-done-circle",
                "title": f"Import complete: {up.get('filename', 'file')}",
                "message": f"{up.get('imported_count', 0)} {up.get('category', 'rows')} imported.",
                "created_at": (ts.isoformat() if isinstance(ts, datetime) else datetime.now(timezone.utc).isoformat()),
                "link": "integrations",
            })

    # Sort: critical first, then warning, then info/success. Most recent first within same severity.
    sev_order = {"critical": 0, "warning": 1, "info": 2, "success": 3}
    notifs.sort(key=lambda n: (sev_order.get(n["severity"], 9), n.get("created_at", "")))
    return notifs


@api_router.get("/notifications")
async def list_notifications(request: Request,
                             session_token: Optional[str] = Cookie(default=None),
                             authorization: Optional[str] = Header(default=None)):
    user = await get_current_user(request, session_token, authorization)
    notifs = await _compute_notifications(user.user_id)
    read_ids = set(
        r["notification_id"] for r in await db.notification_reads.find(
            {"user_id": user.user_id}, {"_id": 0, "notification_id": 1}
        ).to_list(5000)
    )
    for n in notifs:
        n["read"] = n["id"] in read_ids
    unread_count = sum(1 for n in notifs if not n["read"])
    return {"notifications": notifs, "unread_count": unread_count}


@api_router.get("/notifications/unread-count")
async def notifications_unread_count(request: Request,
                                     session_token: Optional[str] = Cookie(default=None),
                                     authorization: Optional[str] = Header(default=None)):
    user = await get_current_user(request, session_token, authorization)
    notifs = await _compute_notifications(user.user_id)
    read_ids = set(
        r["notification_id"] for r in await db.notification_reads.find(
            {"user_id": user.user_id}, {"_id": 0, "notification_id": 1}
        ).to_list(5000)
    )
    return {"unread_count": sum(1 for n in notifs if n["id"] not in read_ids)}


class MarkReadRequest(BaseModel):
    notification_id: str


@api_router.post("/notifications/read")
async def mark_read(req: MarkReadRequest, request: Request,
                    session_token: Optional[str] = Cookie(default=None),
                    authorization: Optional[str] = Header(default=None)):
    user = await get_current_user(request, session_token, authorization)
    await db.notification_reads.update_one(
        {"user_id": user.user_id, "notification_id": req.notification_id},
        {"$set": {
            "user_id": user.user_id,
            "notification_id": req.notification_id,
            "read_at": datetime.now(timezone.utc),
        }},
        upsert=True,
    )
    return {"ok": True}


@api_router.post("/notifications/read-all")
async def mark_all_read(request: Request,
                       session_token: Optional[str] = Cookie(default=None),
                       authorization: Optional[str] = Header(default=None)):
    user = await get_current_user(request, session_token, authorization)
    notifs = await _compute_notifications(user.user_id)
    now = datetime.now(timezone.utc)
    for n in notifs:
        await db.notification_reads.update_one(
            {"user_id": user.user_id, "notification_id": n["id"]},
            {"$set": {"user_id": user.user_id, "notification_id": n["id"], "read_at": now}},
            upsert=True,
        )
    return {"ok": True, "count": len(notifs)}


# =================== END-OF-DAY REPORT ===================
async def _build_eod_report(user_id: str, date_str: Optional[str] = None,
                             branch_id: Optional[str] = None) -> Dict[str, Any]:
    date_str = date_str or datetime.now(timezone.utc).date().isoformat()
    user = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    q = {"user_id": user_id}
    if branch_id and branch_id != "all":
        q["branch_id"] = branch_id

    branches = await db.branches.find({"user_id": user_id}, {"_id": 0}).to_list(50)
    branch_name = {b["branch_id"]: b["name"] for b in branches}

    # Today sales
    sales = await db.sales.find({**q, "date": date_str}, {"_id": 0}).to_list(500)
    total_sales = sum(s["total"] for s in sales)
    total_orders = sum(s["orders"] for s in sales)
    avg_ticket = (total_sales / total_orders) if total_orders else 0
    cash = sum(s["cash"] for s in sales)
    upi = sum(s["upi"] for s in sales)
    card = sum(s["card"] for s in sales)
    tax = sum(s.get("tax", 0) for s in sales)

    # Expenses today
    exps_today = await db.expenses.find({**q, "date": date_str}, {"_id": 0}).to_list(500)
    total_exp_today = sum(e["amount"] for e in exps_today)

    profit_est = total_sales - total_exp_today

    # Branch-wise
    per_branch: Dict[str, Dict[str, Any]] = {}
    for s in sales:
        b = per_branch.setdefault(s["branch_id"], {
            "branch_id": s["branch_id"],
            "branch_name": branch_name.get(s["branch_id"], "Branch"),
            "total": 0, "orders": 0,
        })
        b["total"] += s["total"]
        b["orders"] += s["orders"]
    branches_perf = sorted(per_branch.values(), key=lambda x: -x["total"])

    # Top / Low selling items (lifetime aggregate)
    menu = await db.menu_items.find(q, {"_id": 0}).to_list(1000)
    agg: Dict[str, Dict[str, Any]] = {}
    for m in menu:
        a = agg.setdefault(m["name"], {"name": m["name"], "units_sold": 0, "revenue": 0})
        a["units_sold"] += m["units_sold"]
        a["revenue"] += m["units_sold"] * m["price"]
    items_sorted = sorted(agg.values(), key=lambda x: -x["units_sold"])
    top_items = items_sorted[:5]
    low_items = [x for x in items_sorted if x["units_sold"] > 0][-5:]

    # Staff attendance summary for today
    att = await db.attendance.find({"user_id": user_id, "date": date_str}, {"_id": 0}).to_list(1000)
    att_counts = {"present": 0, "absent": 0, "leave": 0, "half": 0}
    for a in att:
        att_counts[a["status"]] = att_counts.get(a["status"], 0) + 1
    total_emp = await db.employees.count_documents({"user_id": user_id})

    # Low stock items count
    inv = await db.inventory.find({"user_id": user_id}, {"_id": 0}).to_list(2000)
    low_stock = [i for i in inv if i["stock"] < i["min_stock"]]

    return {
        "report_for": date_str,
        "business_name": (user or {}).get("business_name") or (user or {}).get("name") or "Your Business",
        "owner_email": (user or {}).get("email", ""),
        "totals": {
            "total_sales": round(total_sales, 2),
            "total_orders": total_orders,
            "avg_ticket": round(avg_ticket, 2),
            "cash": round(cash, 2),
            "upi": round(upi, 2),
            "card": round(card, 2),
            "tax": round(tax, 2),
            "expenses_today": round(total_exp_today, 2),
            "profit_estimate": round(profit_est, 2),
        },
        "branches": branches_perf,
        "top_items": top_items,
        "low_items": low_items,
        "staff_attendance": {**att_counts, "total_employees": total_emp},
        "low_stock_count": len(low_stock),
        "low_stock_preview": [
            {"name": i["name"], "stock": i["stock"], "min_stock": i["min_stock"], "unit": i["unit"]}
            for i in low_stock[:10]
        ],
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }


@api_router.get("/reports/eod")
async def eod_report(request: Request, date: Optional[str] = None, branch_id: Optional[str] = None,
                     session_token: Optional[str] = Cookie(default=None),
                     authorization: Optional[str] = Header(default=None)):
    user = await get_current_user(request, session_token, authorization)
    return await _build_eod_report(user.user_id, date, branch_id)


@api_router.get("/reports/eod/pdf")
async def eod_report_pdf(request: Request, date: Optional[str] = None, branch_id: Optional[str] = None,
                         session_token: Optional[str] = Cookie(default=None),
                         authorization: Optional[str] = Header(default=None)):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from fastapi.responses import StreamingResponse
    import io as _io_pdf

    user = await get_current_user(request, session_token, authorization)
    report = await _build_eod_report(user.user_id, date, branch_id)

    buf = _io_pdf.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm,
                            leftMargin=1.5 * cm, rightMargin=1.5 * cm, title="EOD Report")

    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Heading1"], textColor=colors.HexColor("#F97316"), fontSize=22, spaceAfter=6)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], textColor=colors.HexColor("#111827"), fontSize=13, spaceBefore=14, spaceAfter=6)
    normal = ParagraphStyle("n", parent=styles["Normal"], fontSize=10, textColor=colors.HexColor("#374151"))
    small = ParagraphStyle("s", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#6B7280"))

    def inr_fmt(n):
        return f"₹ {float(n or 0):,.2f}"

    story = []
    story.append(Paragraph("RestroHub — End of Day Report", h1))
    story.append(Paragraph(f"{report['business_name']} · {report['report_for']}", small))
    story.append(Spacer(1, 6))

    # Totals table
    t = report["totals"]
    totals_data = [
        ["Total Sales", inr_fmt(t["total_sales"]), "Orders", str(t["total_orders"])],
        ["Avg Ticket", inr_fmt(t["avg_ticket"]), "Tax Collected", inr_fmt(t["tax"])],
        ["Cash", inr_fmt(t["cash"]), "UPI", inr_fmt(t["upi"])],
        ["Card", inr_fmt(t["card"]), "Expenses Today", inr_fmt(t["expenses_today"])],
        ["Profit Estimate", inr_fmt(t["profit_estimate"]), "Low Stock Items", str(report["low_stock_count"])],
    ]
    tbl = Table(totals_data, colWidths=[4 * cm, 4.5 * cm, 4 * cm, 4.5 * cm])
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FFF7ED")),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#F97316")),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#FDBA74")),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#111827")),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
    ]))
    story.append(Paragraph("Today at a Glance", h2))
    story.append(tbl)

    # Branches
    if report["branches"]:
        story.append(Paragraph("Branch-wise Performance", h2))
        rows = [["Branch", "Sales", "Orders"]]
        for b in report["branches"]:
            rows.append([b["branch_name"], inr_fmt(b["total"]), str(b["orders"])])
        bt = Table(rows, colWidths=[8 * cm, 5 * cm, 4 * cm])
        bt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111827")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D1D5DB")),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(bt)

    # Top items
    if report["top_items"]:
        story.append(Paragraph("Top Selling Items", h2))
        rows = [["#", "Item", "Units", "Revenue"]]
        for i, it in enumerate(report["top_items"], 1):
            rows.append([str(i), it["name"], str(it["units_sold"]), inr_fmt(it["revenue"])])
        tt = Table(rows, colWidths=[1 * cm, 9 * cm, 3 * cm, 4 * cm])
        tt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F97316")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D1D5DB")),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(tt)

    # Staff attendance
    sa = report["staff_attendance"]
    story.append(Paragraph("Staff Attendance Today", h2))
    sa_rows = [
        ["Total Employees", str(sa["total_employees"])],
        ["Present", str(sa["present"])],
        ["Half-day", str(sa["half"])],
        ["Leave", str(sa["leave"])],
        ["Absent", str(sa["absent"])],
    ]
    sat = Table(sa_rows, colWidths=[8 * cm, 4 * cm])
    sat.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D1D5DB")),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
    ]))
    story.append(sat)

    # Low stock preview
    if report["low_stock_preview"]:
        story.append(Paragraph("Low Stock Items (Reorder Needed)", h2))
        rows = [["Item", "Current", "Minimum"]]
        for i in report["low_stock_preview"]:
            rows.append([i["name"], f"{i['stock']:g} {i['unit']}", f"{i['min_stock']:g} {i['unit']}"])
        lst = Table(rows, colWidths=[8 * cm, 4 * cm, 4 * cm])
        lst.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EF4444")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#FCA5A5")),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(lst)

    story.append(Spacer(1, 18))
    story.append(Paragraph(f"Generated by RestroHub · {report['generated_at'][:19]} UTC", small))
    doc.build(story)
    buf.seek(0)

    filename = f"RestroHub-EOD-{report['report_for']}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


class EodSendRequest(BaseModel):
    date: Optional[str] = None
    branch_id: Optional[str] = None


@api_router.post("/reports/eod/send")
async def eod_report_send(req: EodSendRequest, request: Request,
                          session_token: Optional[str] = Cookie(default=None),
                          authorization: Optional[str] = Header(default=None)):
    """Stub for email delivery. Currently marks a queued job when SendGrid isn't configured."""
    user = await get_current_user(request, session_token, authorization)
    settings = await _get_settings(user.user_id)
    recipients = settings.get("eod_recipient_emails") or []
    has_email_creds = bool(os.environ.get("SENDGRID_API_KEY")) and bool(os.environ.get("SENDER_EMAIL"))
    if not has_email_creds:
        return {
            "ok": False,
            "queued": False,
            "reason": "email_not_configured",
            "message": "SendGrid is not configured yet. Add SENDGRID_API_KEY and SENDER_EMAIL in backend/.env to enable sending.",
            "recipients": recipients,
        }
    if not recipients:
        raise HTTPException(status_code=400, detail="Add at least one recipient email in Settings first.")
    # Email delivery path is disabled until the user provides SendGrid credentials.
    return {
        "ok": False,
        "queued": False,
        "reason": "email_delivery_disabled",
        "message": "Email delivery is disabled in this build. PDF is available via /api/reports/eod/pdf.",
        "recipients": recipients,
    }


app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()


app.include_router(api_router)
